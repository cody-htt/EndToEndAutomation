import openpyxl


def data_generator():
    my_list = [
        ['Tung1', 'tunghuynh1@gmail.com', 'AutoQC1', 'AutoQC1', '123 ABC'],
        ['Tung2', 'tunghuynh2@gmail.com', 'AutoQC2', 'AutoQC2', '456 ABC'],
        ['Tung3', 'tunghuynh3@gmail.com', 'AutoQC3', 'AutoQC3', '789 ABC'],
        ['Tung4', 'tunghuynh4@gmail.com', 'AutoQC4', 'AutoQC4', '434132 ABC'],
        ['Tung5', 'tunghuynh5@gmail.com', 'AutoQC5', 'AutoQC5', '4254989 ABC']
    ]
    return my_list


def excel_generator():
    workbook = openpyxl.load_workbook("F:/9. Python Selenium RobotFW - Automation Project (Study)/Book1.xlsx")
    sheet_1 = workbook['Sheet1']
    rows = sheet_1.max_row
    my_list = []
    for row in range(1, rows+1):
        temp_list = []
        username = sheet_1.cell(row, 1)
        email = sheet_1.cell(row, 2)
        user_password = sheet_1.cell(row, 3)
        repeat_password = sheet_1.cell(row, 4)
        address = sheet_1.cell(row, 5)
        temp_list.insert(0, username.value)
        temp_list.insert(1, email.value)
        temp_list.insert(2, user_password.value)
        temp_list.insert(3, repeat_password.value)
        temp_list.insert(4, address.value)
        my_list.insert(row-1, temp_list)

    print(my_list)
    return my_list
